from openpyxl import Workbook

# Create a new Workbook
wb = Workbook()

# Get the active sheet
ws = wb.active

# Add data to the sheet
ws['A1'] = 'Student Name'
ws['B1'] = 'Maths Marks'
ws['C1'] = 'Science Marks'

ws['A2'] = 'Alice'
ws['B2'] = 85
ws['C2'] = 90

ws['A3'] = 'Bob'
ws['B3'] = 78
ws['C3'] = 88

# Save the workbook
wb.save('student_marks.xlsx')
